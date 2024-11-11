# Name: openpyxl
# Version: 3.1.5
# Summary: A Python library to read/write Excel 2010 xlsx/xlsm files
# pip show openpyxl
import mysql.connector #for connect with mysql
from openpyxl import load_workbook  #for excel related work and load_workbook it's allow you to open and work with 
# existing excel file   
# excel_file="C:\Users\Avinash singh\OneDrive\Desktop\data.xlsx"   wrong currect is //
excel_file = "C:\\Users\\Avinash singh\\OneDrive\\Desktop\\data.xlsx"
workbook = load_workbook(excel_file)
sheet = workbook.active # use for active this excel sheet
data = []
columns = [cell.value for cell in sheet[1]]  # First row is assumed to be the column headers
columns = [f"`{col}`" for col in columns]
# this columns line create list ,this store excel first row header for excel file
# [cell.value for cell in sheet[1]] means "take the value of each cell in the first row and put it in the columns list."

# # print(columns) working
for row in sheet.iter_rows(min_row=2, values_only=True):  # min_row=2 to start reading from the second row
    row_data = dict(zip(columns, row))  
    data.append(row_data)
# iter_rows() is a method provided by openpyxl to iterate through rows in an Excel sheet
# When you call sheet.iter_rows() in Python (using a library like openpyxl), it doesn't immediately give you all the 
# rows of the Excel sheet at once. Instead, it gives you a generator
# min_col=2 tells iter_rows() to start reading from the second column of each row and skip the first column.
# Setting values_only=True ensures that only the values of each cell are retrieved, not the Cell objects themselves.
    # row_data = dict(zip(columns, row))  # Create a dictionary for each row
#   data.append(row_data)
# print(data) working
connection = mysql.connector.connect(
    host='localhost',          # Replace with your host, e.g., 'localhost'
    user='root',       # Replace with your MySQL username
    password='AVInash12@#',   # Replace with your MySQL password
    database='exceldata'    # Replace with your database name
)
cursor = connection.cursor()
# # Step 4: Create table if it does not exist ..................................
Excel_indatabase = ', '.join([f"{col} VARCHAR(255)" for col in columns])  # Use backticks for column names
create_table_query = f"CREATE TABLE IF NOT EXISTS EXceldata ({Excel_indatabase})"
cursor.execute(create_table_query)

# Step 5: Insert Excel data into the MySQL table
for row_data in data:   
    placeholders = ', '.join(['%s'] * len(columns))
    insert_query = f"INSERT INTO EXceldata ({', '.join(columns)}) VALUES ({placeholders})"
    cursor.execute(insert_query, tuple(row_data.values()))

# Step 6: Commit and close the database connection
connection.commit()
cursor.close()
connection.close()

print(create_table_query)
print(insert_query)






